import re
import pandas as pd
import openpyxl
import numpy as np
import calendar
from datetime import datetime
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import re
from openpyxl.worksheet.table import Table, TableStyleInfo
import io


change_month = {
    'Enero':1, 
    'Febrero':2, 
    'Marzo':3, 
    'Abril':4, 
    'Mayo':5, 
    'Jun':6,  
    'Jul':7, 
    'Ago':8, 
    'Set':9, 
    'Oct':10,
    'Nov':11, 
    'Dic':12
}
change_month_TEXT = {
    'ENERO':1, 
    'FEBRERO':2, 
    'MARZO':3, 
    'ABRIL':4, 
    'MAYO':5, 
    'JUNIO':6,  
    'JULIO':7, 
    'AGOSTO':8, 
    'SEPTIEMBRE':9, 
    'OCTUBRE':10,
    'NOVIEMBRE':11, 
    'DICIEMBRE':12
}
def limpiar_kg_exportables(valor):
        # Convertir a string para poder verificar si contiene punto
        valor_str = str(valor)
        
        if "." in valor_str:
            # Si contiene punto, convertir a float, multiplicar por 1000 y convertir a entero
            return float(valor_str) * 1000
        else:
            # Si no contiene punto, solo convertir a entero
            return int(valor_str)
        
def split_if_colon_at_3(texto):
    if isinstance(texto, str) and len(texto) > 2 and texto[2] == ':':
        return [texto[:2], texto[3:].strip()]
    else:
        return [None, texto]
    
def get_month_name(month_number: int) -> str:
   
    if not 1 <= month_number <= 12:
        raise ValueError("El número de mes debe estar entre 1 y 12")
        
    months = {
        1: "ENERO",
        2: "FEBRERO",
        3: "MARZO",
        4: "ABRIL",
        5: "MAYO",
        6: "JUNIO",
        7: "JULIO",
        8: "AGOSTO",
        9: "SETIEMBRE",
        10: "OCTUBRE",
        11: "NOVIEMBRE",
        12: "DICIEMBRE"
    }
    
    return months[month_number]
# Validar y corregir HORA RECEPCION para que siempre sea de la tarde
def corregir_hora_tarde(hora_str):
        if pd.isna(hora_str):
            return hora_str
        match = re.match(r"(\d{2}):(\d{2}):(\d{2})", str(hora_str))
        if not match:
            return hora_str
        h, m, s = map(int, match.groups())
        # Si la hora es menor a 12, sumamos 12 horas para que sea PM
        if h < 12:
            h += 12
            if h == 24:
                h = 12  # 12 PM
        return f"{h:02d}:{m:02d}:{s:02d}"


def create_format_excel(dff: pd.DataFrame, nombre_archivo: str) -> str:
     with pd.ExcelWriter(nombre_archivo, engine="openpyxl") as writer:
        dff.to_excel(writer, index=False, sheet_name="TIEMPOS")
        ws = writer.sheets["TIEMPOS"]

        # Encabezados en negrita y fondo azul claro
        header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
        for col_num, col in enumerate(dff.columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
            cell.fill = header_fill

        # Ajustar ancho de columnas automáticamente
        for i, col in enumerate(dff.columns, 1):
            max_length = max(
                [len(str(cell.value)) if cell.value is not None else 0 for cell in ws[get_column_letter(i)]]
            )
            ws.column_dimensions[get_column_letter(i)].width = max_length + 2

        # Congelar la primera fila
        ws.freeze_panes = "A2"

        # Validar encabezados para tabla de Excel
        columnas_validas = True
        colnames = list(dff.columns)
        if any(pd.isna(col) or str(col).strip() == '' for col in colnames):
            columnas_validas = False
        if len(set(colnames)) != len(colnames):
            columnas_validas = False
        if any(any(c in str(col) for c in ['[', ']', '*', '?', '/', '\\']) for col in colnames):
            columnas_validas = False

        # Crear tabla de Excel real solo si los encabezados son válidos
        if columnas_validas:
            nrows = dff.shape[0] + 1  # +1 por encabezado
            ncols = dff.shape[1]
            last_col = get_column_letter(ncols)
            table_ref = f"A1:{last_col}{nrows}"
            tabla = Table(displayName="TIEMPOS_TABLA", ref=table_ref)
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            tabla.tableStyleInfo = style
            ws.add_table(tabla)
        else:
            print("No se pudo crear la tabla de Excel porque los encabezados no son válidos para Excel. Solo se exportó el formato básico.")

def create_format_excel_in_memory(dff: pd.DataFrame) -> bytes:
    """
    Crea un archivo Excel formateado en memoria y retorna los bytes
    
    Args:
        dff: DataFrame de pandas a formatear
    
    Returns:
        bytes: Contenido del archivo Excel formateado
    """
    # Crear buffer en memoria
    excel_buffer = io.BytesIO()
    
    with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
        dff.to_excel(writer, index=False, sheet_name="TIEMPOS")
        ws = writer.sheets["TIEMPOS"]

        # Encabezados en negrita y fondo azul claro
        header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
        for col_num, col in enumerate(dff.columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
            cell.fill = header_fill

        # Ajustar ancho de columnas automáticamente
        for i, col in enumerate(dff.columns, 1):
            max_length = max(
                [len(str(cell.value)) if cell.value is not None else 0 for cell in ws[get_column_letter(i)]]
            )
            ws.column_dimensions[get_column_letter(i)].width = max_length + 2

        # Congelar la primera fila
        ws.freeze_panes = "A2"

        # Validar encabezados para tabla de Excel
        columnas_validas = True
        colnames = list(dff.columns)
        if any(pd.isna(col) or str(col).strip() == '' for col in colnames):
            columnas_validas = False
        if len(set(colnames)) != len(colnames):
            columnas_validas = False
        if any(any(c in str(col) for c in ['[', ']', '*', '?', '/', '\\']) for col in colnames):
            columnas_validas = False

        # Crear tabla de Excel real solo si los encabezados son válidos
        if columnas_validas:
            nrows = dff.shape[0] + 1  # +1 por encabezado
            ncols = dff.shape[1]
            last_col = get_column_letter(ncols)
            table_ref = f"A1:{last_col}{nrows}"
            tabla = Table(displayName="TIEMPOS_TABLA", ref=table_ref)
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            tabla.tableStyleInfo = style
            ws.add_table(tabla)
        else:
            print("⚠️  No se pudo crear la tabla de Excel porque los encabezados no son válidos. Solo se aplicó el formato básico.")
    
    # Obtener los bytes del archivo Excel formateado
    excel_data = excel_buffer.getvalue()
    excel_buffer.close()
    
    return excel_data

def get_download_url_by_name(json_data, name,):
    """
    Busca en el JSON un archivo por su nombre y retorna su downloadUrl
    
    Args:
        json_data (list): Lista de diccionarios con información de archivos
        name (str): Nombre del archivo a buscar
    
    Returns:
        str: URL de descarga del archivo encontrado, o None si no se encuentra
    """
    for item in json_data:
        if item.get('name') == name:
            return item.get('@microsoft.graph.downloadUrl')
        
def structure_planilla_historica_like_estimate(df_planilla_historica):
    """
    Estructura la planilla histórica igual que estimate_current_planilla_by_previous:
    Para cada mes y proyecto, distribuye el costo total entre los días laborables (lunes a viernes) del mes.

    Args:
        df_planilla_historica (pd.DataFrame): DataFrame con columnas ['Mes', 'DESCRIPCION PROYECTO', 'Costos']

    Returns:
        pd.DataFrame: DataFrame con columnas ['DESCRIPCION PROYECTO', 'FECHA', 'TOTAL', 'MES_ACTUAL', 'AÑO_ACTUAL']
    """
    if not np.issubdtype(df_planilla_historica['Mes'].dtype, np.datetime64):
        df_planilla_historica = df_planilla_historica.copy()
        df_planilla_historica['Mes'] = pd.to_datetime(df_planilla_historica['Mes'])

    resultados = []
    for (año, mes), df_mes in df_planilla_historica.groupby([df_planilla_historica['Mes'].dt.year, df_planilla_historica['Mes'].dt.month]):
        # Solo días laborables (lunes=0, ..., viernes=4)
        dias_mes = pd.date_range(start=datetime(año, mes, 1), end=datetime(año, mes, pd.Period(f'{año}-{mes:02d}').days_in_month))
        dias_laborables = dias_mes[dias_mes.weekday < 5]  # lunes a viernes
        df_grouped = df_mes.groupby('DESCRIPCION PROYECTO', as_index=False)['Costos'].sum()
        for _, row in df_grouped.iterrows():
            costo_diario = row['Costos'] / len(dias_laborables) if len(dias_laborables) > 0 else 0
            for fecha in dias_laborables:
                resultados.append({
                    'DESCRIPCION PROYECTO': row['DESCRIPCION PROYECTO'],
                    'FECHA': fecha,
                    'TOTAL': costo_diario,
                })
    
    return pd.DataFrame(resultados)

def estimate_current_planilla_by_previous(df_planilla_historica):
    """
    Calcula la planilla "actual" (mes más reciente sin datos) usando la planilla del mes anterior,
    agrupando por proyecto y distribuyendo el costo total entre los días laborables (lunes a viernes) del mes actual.
    Si existe la planilla del mes actual, se usa esa. Si no, se usa la del mes anterior.

    Args:
        df_planilla_historica (pd.DataFrame): DataFrame con columnas ['Mes', 'DESCRIPCION PROYECTO', 'Costos']
            donde 'Mes' es tipo datetime o string 'YYYY-MM'.

    Returns:
        pd.DataFrame: DataFrame con columnas ['DESCRIPCION PROYECTO', 'DIA', 'COSTO_DIARIO', 'MES_ACTUAL', 'AÑO_ACTUAL']
        donde 'DIA' es una fecha completa (datetime)
    """
    # Normalizar columna 'Mes' a datetime
    if not np.issubdtype(df_planilla_historica['Mes'].dtype, np.datetime64):
        df_planilla_historica = df_planilla_historica.copy()
        df_planilla_historica['Mes'] = pd.to_datetime(df_planilla_historica['Mes'])

    # Encontrar el mes más reciente en la planilla histórica
    max_mes = df_planilla_historica['Mes'].max()
    
    año_max = max_mes.year
    mes_max = max_mes.month
    
    # Calcular el mes "actual" (el siguiente al más reciente en la planilla)
    if mes_max == 12:
        año_actual = año_max + 1
        mes_actual = 1
    else:
        año_actual = año_max
        mes_actual = mes_max + 1

    # Verificar si ya existe la planilla del mes actual
    existe_actual = (
        (df_planilla_historica['Mes'].dt.year == año_actual) &
        (df_planilla_historica['Mes'].dt.month == mes_actual)
    ).any()

    if existe_actual:
        # Usar la planilla del mes actual
        df_mes = df_planilla_historica[
            (df_planilla_historica['Mes'].dt.year == año_actual) &
            (df_planilla_historica['Mes'].dt.month == mes_actual)
        ]
        # Solo días laborables (lunes=0, ..., viernes=4)
        dias_mes_actual = pd.date_range(start=datetime(año_actual, mes_actual, 1), end=datetime(año_actual, mes_actual, pd.Period(f'{año_actual}-{mes_actual:02d}').days_in_month))
        dias_laborables = dias_mes_actual[dias_mes_actual.weekday < 5]
    else:
        # Usar la planilla del mes anterior
        df_mes = df_planilla_historica[
            (df_planilla_historica['Mes'].dt.year == año_max) &
            (df_planilla_historica['Mes'].dt.month == mes_max)
        ]
        # El mes actual es el siguiente al más reciente
        dias_mes_actual = pd.date_range(start=datetime(año_actual, mes_actual, 1), end=datetime(año_actual, mes_actual, pd.Period(f'{año_actual}-{mes_actual:02d}').days_in_month))
        dias_laborables = dias_mes_actual[dias_mes_actual.weekday < 5]

    # Agrupar por proyecto y sumar costos
    df_grouped = df_mes.groupby('DESCRIPCION PROYECTO', as_index=False)['Costos'].sum()
    num_dias_laborables = len(dias_laborables)
    df_grouped['COSTO_DIARIO'] = df_grouped['Costos'] / num_dias_laborables if num_dias_laborables > 0 else 0

    # Expandir a cada día laborable del mes actual como fecha completa
    df_result = pd.DataFrame([
        {
            'DESCRIPCION PROYECTO': row['DESCRIPCION PROYECTO'],
            'FECHA': fecha,
            'TOTAL': row['COSTO_DIARIO'],
            'MES_ACTUAL': mes_actual,
            'AÑO_ACTUAL': año_actual
        }
        for _, row in df_grouped.iterrows() for fecha in dias_laborables
    ])
    df_result = df_result[["DESCRIPCION PROYECTO","FECHA","TOTAL"]]
    return df_result

def generate_date_options_dataframe(start_year=2024, start_month=8):
    """
    Genera un DataFrame con opciones de YEAR, MONTH y WEEK desde agosto 2024 hasta la fecha actual.
    
    Args:
        start_year (int): Año de inicio (por defecto 2024)
        start_month (int): Mes de inicio (por defecto 8 = agosto)
    
    Returns:
        dict: Diccionario con tres DataFrames:
            - 'years': DataFrame con columnas ['value', 'label'] para años disponibles
            - 'months': DataFrame con columnas ['value', 'label', 'year'] para meses disponibles por año  
            - 'weeks': DataFrame con columnas ['value', 'label', 'year', 'month'] para semanas
    """
    from datetime import datetime, timedelta
    import pandas as pd
    
    current_date = datetime.now()
    start_date = datetime(start_year, start_month, 1)
    
    # Si la fecha de inicio es posterior a la actual, usar la fecha actual
    if start_date > current_date:
        start_date = datetime(current_date.year, current_date.month, 1)
    
    # Generar años disponibles
    years_data = []
    available_years = set()
    
    # Iterar desde la fecha de inicio hasta la fecha actual
    temp_date = start_date
    while temp_date <= current_date:
        available_years.add(temp_date.year)
        # Avanzar al siguiente mes
        if temp_date.month == 12:
            temp_date = datetime(temp_date.year + 1, 1, 1)
        else:
            temp_date = datetime(temp_date.year, temp_date.month + 1, 1)
    
    for year in sorted(available_years):
        years_data.append({
            'value': str(year),
            'label': str(year)
        })
    
    # Generar meses disponibles por año
    months_data = []
    month_names = [
        "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
    ]
    
    for year in sorted(available_years):
        # Determinar rango de meses para este año
        if year == start_year:
            start_month_for_year = start_month
        else:
            start_month_for_year = 1
            
        if year == current_date.year:
            end_month_for_year = current_date.month
        else:
            end_month_for_year = 12
        
        for month in range(start_month_for_year, end_month_for_year + 1):
            months_data.append({
                'value': str(month),
                'label': month_names[month - 1],
                'year': year
            })
    
    # Generar semanas para cada año y mes disponible
    weeks_data = []
    
    for year in sorted(available_years):
        # Determinar rango de meses para este año
        if year == start_year:
            start_month_for_year = start_month
        else:
            start_month_for_year = 1
            
        if year == current_date.year:
            end_month_for_year = current_date.month
        else:
            end_month_for_year = 12
        
        for month in range(start_month_for_year, end_month_for_year + 1):
            # Obtener el primer día del mes
            first_day = datetime(year, month, 1)
            
            # Obtener el último día del mes
            if month == 12:
                last_day = datetime(year + 1, 1, 1) - timedelta(days=1)
            else:
                last_day = datetime(year, month + 1, 1) - timedelta(days=1)
            
            # Si es el mes actual, limitar hasta la fecha actual
            if year == current_date.year and month == current_date.month:
                last_day = min(last_day, current_date)
            
            # Encontrar el primer lunes del mes o antes
            start_week = first_day - timedelta(days=first_day.weekday())
            
            week_number = 1
            current_week_start = start_week
            
            while current_week_start <= last_day:
                # Calcular el final de la semana
                current_week_end = current_week_start + timedelta(days=6)
                
                # Verificar si la semana tiene días en el mes actual
                if (current_week_start.month == month and current_week_start.year == year) or \
                   (current_week_end.month == month and current_week_end.year == year) or \
                   (current_week_start <= first_day <= current_week_end):
                    
                    week_label = f"Semana {week_number} ({current_week_start.strftime('%d/%m')} - {current_week_end.strftime('%d/%m')})"
                    
                    weeks_data.append({
                        'value': f"{year}-{month:02d}-W{week_number}",
                        'label': week_label,
                        'year': year,
                        'month': month,
                        'week_number': week_number,
                        'start_date': current_week_start.strftime('%Y-%m-%d'),
                        'end_date': current_week_end.strftime('%Y-%m-%d')
                    })
                    
                    week_number += 1
                
                # Avanzar a la siguiente semana
                current_week_start += timedelta(weeks=1)
    
    return {
        'years': pd.DataFrame(years_data),
        'months': pd.DataFrame(months_data), 
        'weeks': pd.DataFrame(weeks_data)
    }

def get_current_date_info(min_year=2024, min_month=8):
    """
    Obtiene información de la fecha actual en formato útil para los selectores.
    Si la fecha actual es anterior al mínimo especificado, usa el mínimo.
    
    Args:
        min_year (int): Año mínimo permitido (por defecto 2024)
        min_month (int): Mes mínimo permitido (por defecto 8 = agosto)
    
    Returns:
        dict: Diccionario con información de la fecha actual:
            - 'current_year': Año actual como string
            - 'current_month': Mes actual como string (número)
            - 'current_month_name': Nombre del mes actual
            - 'current_week': Identificador de la semana actual
    """
    from datetime import datetime
    
    current_date = datetime.now()
    min_date = datetime(min_year, min_month, 1)
    
    # Si la fecha actual es anterior al mínimo, usar el mínimo
    if current_date < min_date:
        effective_date = min_date
    else:
        effective_date = current_date
    
    # Calcular el número de semana del mes
    first_day_of_month = datetime(effective_date.year, effective_date.month, 1)
    days_since_start = (effective_date - first_day_of_month).days
    week_of_month = (days_since_start // 7) + 1
    
    month_names = [
        "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
    ]
    
    return {
        'current_year': str(effective_date.year),
        'current_month': str(effective_date.month),
        'current_month_name': month_names[effective_date.month - 1],
        'current_week': f"{effective_date.year}-{effective_date.month:02d}-W{week_of_month}",
        'current_date': effective_date.strftime('%Y-%m-%d')
    }

def generate_list_month(start_year, start_month):
    """
    Genera un DataFrame con todas las fechas desde el año/mes de inicio hasta la fecha actual,
    luego extrae año, mes y semana, y agrupa por estas columnas.
    
    Args:
        start_year (int): Año de inicio
        start_month (int): Mes de inicio (1-12)
    
    Returns:
        pd.DataFrame: DataFrame agrupado con columnas ['YEAR', 'MES', 'SEMANA']
            - YEAR: Año extraído de la fecha
            - MES: Nombre del mes en español
            - SEMANA: Número de semana por mes (se reinicia en cada mes)
    """
    from datetime import datetime, timedelta
    import pandas as pd
    
    # Fecha actual
    current_date = datetime.now()
    
    # Fecha de inicio (primer día del mes de inicio)
    start_date = datetime(start_year, start_month, 1)
    
    # Si la fecha de inicio es posterior a la actual, usar la fecha actual
    if start_date > current_date:
        start_date = datetime(current_date.year, current_date.month, 1)
        start_year = current_date.year
        start_month = current_date.month
    
    # Nombres de meses en español
    month_names = [
        "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
    ]
    
    # Generar todas las fechas desde la fecha de inicio hasta la fecha actual
    date_list = []
    current_date_iter = start_date
    
    while current_date_iter <= current_date:
        date_list.append(current_date_iter)
        current_date_iter += timedelta(days=1)
    
    # Crear DataFrame con columna FECHA
    df = pd.DataFrame({'FECHA': date_list})
    
    # Extraer año, mes y semana de la columna FECHA
    df['YEAR'] = df['FECHA'].dt.year
    df['MES'] = df['FECHA'].dt.month
    df['MES_TEXT'] = df['FECHA'].dt.month.map(lambda x: month_names[x - 1])
    
    
    df['SEMANA'] = df['FECHA'].dt.isocalendar().week 
    # Agrupar por YEAR, MES y SEMANA
    grouped_df = df.groupby(['YEAR', 'MES','MES_TEXT', 'SEMANA']).size().reset_index(name='count')
    
    # Eliminar la columna count ya que solo necesitamos las columnas de agrupación
    grouped_df = grouped_df[['YEAR', 'MES', 'MES_TEXT', 'SEMANA']]  
    
    return grouped_df

def dataframe_filtro(values=[], columns_df=[]):
        """
        Genera una query de filtrado para pandas DataFrame
        
        Args:
            values: Lista de valores para filtrar (puede contener None)
            columns_df: Lista de nombres de columnas correspondientes
            
        Returns:
            str: Query string para pandas.DataFrame.query() o cadena vacía
        """
        if not values or not columns_df:
            return ""
            
        query = ""
        try:
            for value, col in zip(values, columns_df):
                if value is not None:
                    if isinstance(value, int):
                        text = f"`{col}` == {value}"
                    elif isinstance(value, str):
                        text = f"`{col}` == '{value}'"
                    elif isinstance(value, list) and value:  # Lista no vacía
                        # Convertir lista de strings a integers si es necesario
                        if col in ['SEMANA', 'Mes']:
                            try:
                                value = [int(v) for v in value if v is not None]
                                if not value:  # Si la lista queda vacía, saltar
                                    continue
                            except (ValueError, TypeError):
                                print(f"⚠️ Error convirtiendo valores de {col}: {value}")
                                continue
                        text = f"`{col}` in {value}"
                    else:
                        continue
                    query += text + " and "
        except Exception as e:
            print(f"❌ Error en dataframe_filtro: {e}")
            return ""
        
        return query[:-5] if query else ""